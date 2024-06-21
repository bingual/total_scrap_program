import asyncio
import logging
import os
import re
import sys
import traceback
from io import BytesIO
from pathlib import Path
from typing import Union, List, Tuple

import aiofiles
import aiohttp
import nest_asyncio
import openpyxl
import pandas as pd
from PIL import Image
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles.fills import fills
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from tqdm.asyncio import tqdm
from datetime import datetime
from typing_extensions import Dict


if getattr(sys, "frozen", False):
    # test.exe로 실행한 경우,test.exe를 보관한 디렉토리의 full path를 취득
    BASE_DIR = Path(sys.executable).parent
else:
    # python test.py로 실행한 경우,test.py를 보관한 디렉토리의 full path를 취득
    BASE_DIR = Path(__file__).resolve().parent

os.chdir(BASE_DIR)

DEFAULT_DIR_NAME = "종합 상품"
ILLEGAL_CHAR_PATTERN = r"[\x00\x0B\x0C]"  # 제어 문자 정규식


def setup_asyncio() -> None:
    nest_asyncio.apply()
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())


def setup_datetime(date_format: str = "%Y-%m-%d_%H_%M") -> str:
    now = datetime.now()
    timestamp = now.strftime(date_format)
    return timestamp


def setup_logging() -> None:
    output_path = BASE_DIR / "logs"
    output_path.mkdir(parents=True, exist_ok=True)

    current_date = setup_datetime("%Y-%m-%d")
    log_filename = f"log_{current_date}.log"
    log_file = output_path / log_filename

    logging.basicConfig(
        filename=log_file,
        level=logging.WARNING,
        format="[%(asctime)s] [%(levelname)s %(filename)s:%(lineno)s] >>\n%(message)s",
        encoding="utf-8-sig",
    )


async def get_logger() -> logging.Logger:
    return logging.getLogger(__name__)


async def get_error_message() -> str:
    return traceback.format_exc()


async def convert_decimal(target: str) -> Union[str, int]:
    if not isinstance(target, str):
        return target
    else:
        target = re.sub(r"[^0-9]", "", target)
        return int(target)


async def convert_string(target: str) -> str:
    if not isinstance(target, str):
        return target
    else:
        target = re.sub(r"[^a-zA-Z가-힣]", "", target)
        return target


async def convert_model(target: str) -> str:
    if not isinstance(target, str):
        return target
    else:
        target = re.sub(r"[가-힣:]", "", target)
        return target


async def calculate_dimension(worksheet: Worksheet) -> None:
    try:
        for column_cells in worksheet.iter_cols():
            length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = (length + 2) * 1.2  # 조정된 폭 계산
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    except Exception as e:
        message = f"엑셀 폭 조정 중에 예외 발생: \n{await get_error_message()}"
        logger = await get_logger()
        logger.error(message)
        print(message)
        raise e


async def cell_pattern_fill(
    df: pd.DataFrame,
    worksheet: Worksheet,
    head_fill_color: str = "4472C4",
    head_font_color: str = "FFFFFF",
    body_fill_color: str = "D9E1F2",
    body_font_color: str = "000000",
    head_border_color: str = "2E5C99",
    body_border_color: str = "B4C6E7",
    fill_type: fills = "solid",
) -> None:
    try:
        # Define border styles
        thin_border_head = Border(
            left=Side(border_style="thin", color=head_border_color),
            right=Side(border_style="thin", color=head_border_color),
            top=Side(border_style="thin", color=head_border_color),
            bottom=Side(border_style="thin", color=head_border_color),
        )
        thin_border_body = Border(
            left=Side(border_style="thin", color=body_border_color),
            right=Side(border_style="thin", color=body_border_color),
            top=Side(border_style="thin", color=body_border_color),
            bottom=Side(border_style="thin", color=body_border_color),
        )

        # Set header row style
        for row in worksheet.iter_rows(
            min_row=1, max_row=1, min_col=1, max_col=df.shape[1]
        ):
            for cell in row:
                cell.fill = PatternFill(
                    start_color=head_fill_color,
                    end_color=head_fill_color,
                    fill_type=fill_type,
                )
                cell.font = Font(color=head_font_color, bold=True)
                cell.border = thin_border_head

        # Set body row style
        for i, row in enumerate(
            worksheet.iter_rows(
                min_row=2, max_row=worksheet.max_row, min_col=1, max_col=df.shape[1]
            )
        ):
            for cell in row:
                if i % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=body_fill_color,
                        end_color=body_fill_color,
                        fill_type=fill_type,
                    )
                    cell.font = Font(color=body_font_color)
                    cell.border = thin_border_body

    except Exception as e:
        message = f"엑셀 서식 지정 중에 예외 발생: \n{await get_error_message()}"
        logger = await get_logger()
        logger.error(message)
        print(message)
        raise e


# 데이터 정제 함수
async def clean_data(data):
    if isinstance(data, str):
        return re.sub(ILLEGAL_CHAR_PATTERN, "", data)
    elif isinstance(data, dict):
        return {key: await clean_data(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [await clean_data(item) for item in data]
    else:
        return data


async def create_xlsx_file(
    data: Union[Dict, List],
    file_name: str = DEFAULT_DIR_NAME,
    sheet_name: str = DEFAULT_DIR_NAME,
) -> BytesIO:

    df = pd.json_normalize(await clean_data(data))

    io = BytesIO()
    io.name = file_name
    try:
        writer = pd.ExcelWriter(io, engine="openpyxl")  # noqa
        df.to_excel(
            writer,
            index=False,
            engine="openpyxl",
            sheet_name=sheet_name,
        )
        workbook = writer.book
        worksheet = workbook.active

        tasks = [
            calculate_dimension(worksheet),
            cell_pattern_fill(df, worksheet),
        ]
        await tqdm.gather(*tasks, desc=f"{file_name} 엑셀 파일 생성중")
        writer._save()  # noqa

    except Exception as e:
        message = f"엑셀 생성 중에 예외 발생: \n{await get_error_message()}"
        logger = await get_logger()
        logger.error(message)
        print(message)
        raise e

    io.seek(0)
    return io


async def save_to_xlsx(
    xlsx_file: BytesIO,
    output_path=BASE_DIR / "스크랩 결과" / "엑셀",
    dirname: str = DEFAULT_DIR_NAME,
) -> None:
    output_path = output_path / dirname
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = setup_datetime("%Y-%m-%d_%H_%M")

    filename = f"{xlsx_file.name}_{timestamp}"
    extension = ".xlsx"

    file_path = output_path / (filename + extension)

    try:
        async with aiofiles.open(file_path, "wb") as f:
            await f.write(xlsx_file.getvalue())

    except Exception as e:
        message = (
            f"엑셀 파일 저장 중에 예외 발생: '{filename}'\n{await get_error_message()}"
        )
        logger = await get_logger()
        logger.error(message)
        print(message)
        raise e


async def download_and_save_image(
    image_url: str,
    output_path: Path,
    filename: str,
    target_size: Tuple[int, int],
) -> None:
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/91.0.4472.124 Safari/537.36"
    }
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(image_url, headers=headers) as resp:
                if resp.status == 200:
                    image_data = BytesIO(await resp.read())
                    image = Image.open(image_data)
                    image = image.resize(target_size)

                    output_image_data = BytesIO()
                    image.save(output_image_data, format="PNG")
                    output_image_data.seek(0)

                    async with aiofiles.open(output_path / filename, mode="wb") as f:
                        await f.write(output_image_data.read())
                else:
                    message = f"이미지 다운로드 실패: '{filename}', '{image_url}', 상태코드: '{resp.status}'"
                    logger = await get_logger()
                    logger.error(message)
                    print(message)

    except Exception as e:
        message = f"이미지 저장 중 예외 발생: '{filename}'\n{await get_error_message()}"
        logger = await get_logger()
        logger.error(message)
        print(message)
        # raise e


async def download_images(
    image_urls: Union[List[str], List[List[str]]],
    dirname: str = DEFAULT_DIR_NAME,
    start_no: int = 1,
    target_size: Tuple[int, int] = (800, 800),
) -> None:

    if image_urls:
        timestamp = setup_datetime("%Y-%m-%d_%H_%M")

        output_path = BASE_DIR / "스크랩 결과" / "이미지" / dirname / f"{timestamp}"
        output_path.mkdir(parents=True, exist_ok=True)
        extension = ".png"

        for i, image_url in tqdm(
            enumerate(image_urls),
            total=len(image_urls),
            desc=f" 이미지 다운로드 중",
        ):
            if isinstance(image_url, list):
                for j, url in enumerate(image_url):
                    await download_and_save_image(
                        image_url=url,
                        output_path=output_path,
                        filename=f"{start_no + i}_{j + 1}{extension}",
                        target_size=target_size,
                    )
            else:
                await download_and_save_image(
                    image_url=image_url,
                    output_path=output_path,
                    filename=f"{start_no + i}{extension}",
                    target_size=target_size,
                )


async def read_data_info_excel_and_download_images(
    file_path: str,
    sheet_name: str = DEFAULT_DIR_NAME,
    target_size: Tuple[int, int] = (800, 800),
) -> None:

    df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

    if not df.empty:
        timestamp = setup_datetime("%Y-%m-%d_%H_%M")

        output_path = BASE_DIR / "스크랩 이미지 업데이트" / "이미지" / timestamp
        output_path.mkdir(parents=True, exist_ok=True)
        extension = ".png"

        for _, row in tqdm(
            df.iterrows(), total=len(df), desc=f"'{sheet_name}' 이미지 다운로드 중"
        ):

            image_src = row["이미지소스"]
            image_src = image_src.split(";\n")
            if isinstance(image_src, list) and len(image_src) > 1:
                for i, src in enumerate(image_src):
                    await download_and_save_image(
                        image_url=src,
                        output_path=output_path,
                        filename=f"{row['상품번호']}_{i + 1}{extension}",
                        target_size=target_size,
                    )
            else:
                await download_and_save_image(
                    image_url=row["이미지소스"],
                    output_path=output_path,
                    filename=f"{row['상품번호']}{extension}",
                    target_size=target_size,
                )


# 거의 사용 안해서 함수 빼놨음
async def update_image_sources(
    file_path: str,
    image_num_list: List[int],
    sheet_name: str = DEFAULT_DIR_NAME,
    target_size: Tuple[int, int] = (800, 800),
) -> None:

    # 기존 코드
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

    if not df.empty:
        timestamp = setup_datetime("%Y-%m-%d_%H_%M")

        output_path = BASE_DIR / "스크랩 이미지 업데이트" / "이미지" / timestamp
        output_path.mkdir(parents=True, exist_ok=True)
        extension = ".png"

        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        for idx, row in tqdm(
            df.iterrows(), total=len(df), desc=f"'{sheet_name}' 이미지 다운로드 중"
        ):
            image_src = row["이미지소스"]
            image_src = image_src.split(",")
            if isinstance(image_src, list) and image_num_list:

                image_num = image_num_list[idx]  # noqa
                image_src = image_src[image_num]

                # 엑셀 파일 업데이트
                cell = sheet.cell(
                    row=idx + 2, column=df.columns.get_loc("이미지소스") + 1  # noqa
                )
                cell.value = image_src

                await download_and_save_image(
                    image_url=image_src,
                    output_path=output_path,
                    filename=f"{row['상품번호']}{extension}",
                    target_size=target_size,
                )

        # 엑셀 파일 저장
        workbook.save(file_path)
